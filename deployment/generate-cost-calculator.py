"""
SOCA Platform — Capacity & Cost Calculator Generator
Generates soca-capacity-cost-calculator.xlsx
Run: python3 generate-cost-calculator.py
"""
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint

# ── Color palette ──────────────────────────────────────────────────────────────
C_HEADER_BG   = "1E3A5F"   # dark navy
C_HEADER_FG   = "FFFFFF"
C_SECTION_BG  = "2E6DA4"   # blue
C_SECTION_FG  = "FFFFFF"
C_INPUT_BG    = "EBF3FB"   # light blue
C_CALC_BG     = "F0F0F0"   # light grey (read-only)
C_TOTAL_BG    = "FFF2CC"   # yellow highlight
C_WARN_BG     = "FFE0B2"   # orange warning
C_GREEN_BG    = "E8F5E9"
C_BORDER      = "BDBDBD"

# ── Style helpers ──────────────────────────────────────────────────────────────
def hdr_font(bold=True, size=11, color="FFFFFF"):
    return Font(name="Calibri", bold=bold, size=size, color=color)

def body_font(bold=False, size=10, color="000000"):
    return Font(name="Calibri", bold=bold, size=size, color=color)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def border_all():
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def border_bottom():
    s = Side(style="medium", color="000000")
    return Border(bottom=s)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def style_header(cell, text, bg=C_HEADER_BG, size=12):
    cell.value = text
    cell.font = Font(name="Calibri", bold=True, size=size, color="FFFFFF")
    cell.fill = fill(bg)
    cell.alignment = align("center")
    cell.border = border_all()

def style_section(cell, text):
    cell.value = text
    cell.font = Font(name="Calibri", bold=True, size=10, color=C_SECTION_FG)
    cell.fill = fill(C_SECTION_BG)
    cell.alignment = align("left")
    cell.border = border_all()

def style_label(cell, text):
    cell.value = text
    cell.font = body_font()
    cell.alignment = align("left")
    cell.border = border_all()

def style_input(cell, value=None, fmt=None):
    if value is not None:
        cell.value = value
    cell.font = Font(name="Calibri", bold=True, size=10, color="1A237E")
    cell.fill = fill(C_INPUT_BG)
    cell.alignment = align("center")
    cell.border = border_all()
    if fmt:
        cell.number_format = fmt

def style_formula(cell, formula, fmt=None):
    cell.value = formula
    cell.font = body_font(color="333333")
    cell.fill = fill(C_CALC_BG)
    cell.alignment = align("center")
    cell.border = border_all()
    if fmt:
        cell.number_format = fmt

def style_total(cell, formula_or_value, fmt=None, bold=True):
    cell.value = formula_or_value
    cell.font = Font(name="Calibri", bold=bold, size=10, color="000000")
    cell.fill = fill(C_TOTAL_BG)
    cell.alignment = align("center")
    cell.border = border_all()
    if fmt:
        cell.number_format = fmt

def style_note(cell, text):
    cell.value = text
    cell.font = Font(name="Calibri", italic=True, size=9, color="757575")
    cell.alignment = align("left", wrap=True)

USD = '"$"#,##0.00'
USD0 = '"$"#,##0'
NUM = '#,##0'
NUM2 = '#,##0.00'
PCT = '0.0%'
GB  = '#,##0.000 "GB"'

# ══════════════════════════════════════════════════════════════════════════════
# BUILD WORKBOOK
# ══════════════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()

# ── Sheet names ────────────────────────────────────────────────────────────────
sh_input   = wb.active;       sh_input.title   = "📥 Inputs"
sh_usage   = wb.create_sheet("📊 Usage & Capacity")
sh_cost    = wb.create_sheet("💰 Cost Breakdown")
sh_infra   = wb.create_sheet("🖥 Infrastructure")
sh_summary = wb.create_sheet("📋 Summary")

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — INPUTS
# ══════════════════════════════════════════════════════════════════════════════
ws = sh_input
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 38
ws.column_dimensions["C"].width = 18
ws.column_dimensions["D"].width = 22
ws.column_dimensions["E"].width = 38

# Title
ws.merge_cells("A1:E1")
style_header(ws["A1"], "SOCA Platform — Capacity & Cost Calculator", size=14)
ws.row_dimensions[1].height = 30

ws.merge_cells("A2:E2")
ws["A2"].value = "Fill in the BLUE cells. All other cells are calculated automatically."
ws["A2"].font = Font(name="Calibri", italic=True, size=10, color="555555")
ws["A2"].alignment = align("center")
ws.row_dimensions[2].height = 18

# ── Section: Edge Deployment ───────────────────────────────────────────────────
ws.merge_cells("A4:E4")
style_section(ws["A4"], "  EDGE DEPLOYMENT  (soca-dashboard + soca-engine per device)")
ws.row_dimensions[4].height = 20

headers = ["Parameter", "Value", "Unit", "Notes"]
for i, h in enumerate(headers, 2):
    c = ws.cell(5, i, h)
    c.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    c.fill = fill("37474F")
    c.alignment = align("center")
    c.border = border_all()
ws.row_dimensions[5].height = 18

rows_edge = [
    ("Number of Edge Devices",         5,    "devices",     "Each device runs soca-dashboard + soca-engine"),
    ("Cameras per Edge Device",        4,    "cameras",     "IP cameras connected to each edge"),
    ("Total Cameras",               None,    "cameras",     "=C6*C7 — auto calculated"),
    ("Detection Events per Camera",    6,    "events/hour", "Avg detections triggered per camera per hour"),
    ("Active Hours per Day",          18,    "hours/day",   "Camera operational hours (0–24)"),
    ("Snapshot Size per Alert",       150,   "KB",          "Typical JPEG snapshot ~50–300 KB"),
    ("GCS Path Prefix per Edge",        1,   "prefix",      "1 prefix per edge (e.g. edge-jakarta-01/)"),
]
for r, (label, val, unit, note) in enumerate(rows_edge, 6):
    style_label(ws.cell(r, 2), label)
    if val is None:
        style_formula(ws.cell(r, 3), "=C6*C7", NUM)
    else:
        style_input(ws.cell(r, 3), val, NUM if isinstance(val, int) else NUM2)
    style_label(ws.cell(r, 4), unit)
    style_note(ws.cell(r, 5), note)

# ── Section: Cloud Deployment ──────────────────────────────────────────────────
ws.merge_cells("A14:E14")
style_section(ws["A14"], "  CLOUD DEPLOYMENT  (soca-control + soca-service)")
ws.row_dimensions[14].height = 20

rows_cloud = [
    ("Cloud Region",            "asia-southeast1",  "",      "GCP region for Pub/Sub, GCS, Cloud Run"),
    ("Ingest API Calls/Alert",   2,                 "calls", "HTTP calls from soca-service → soca-control per alert"),
]
for r, (label, val, unit, note) in enumerate(rows_cloud, 15):
    style_label(ws.cell(r, 2), label)
    style_input(ws.cell(r, 3), val)
    style_label(ws.cell(r, 4), unit)
    style_note(ws.cell(r, 5), note)

# ── Section: Retention & Time ──────────────────────────────────────────────────
ws.merge_cells("A18:E18")
style_section(ws["A18"], "  RETENTION & TIME PARAMETERS")
ws.row_dimensions[18].height = 20

rows_time = [
    ("Days per Month",         30,   "days",   "Used for monthly cost projection"),
    ("GCS Snapshot Retention", 90,   "days",   "How long snapshots kept in GCS Standard tier"),
    ("Pub/Sub Message Size",  1.5,   "KB",     "Pub/Sub payload overhead (JSON detection data)"),
    ("Pub/Sub Ack Deadline",  600,   "sec",    "Max time before message redelivery"),
]
for r, (label, val, unit, note) in enumerate(rows_time, 19):
    style_label(ws.cell(r, 2), label)
    style_input(ws.cell(r, 3), val, NUM2 if isinstance(val, float) else NUM)
    style_label(ws.cell(r, 4), unit)
    style_note(ws.cell(r, 5), note)

# ── Section: GCP Pricing Reference ────────────────────────────────────────────
ws.merge_cells("A24:E24")
style_section(ws["A24"], "  GCP PRICING REFERENCE  (edit to update pricing)")
ws.row_dimensions[24].height = 20

pricing = [
    ("Pub/Sub — per GB throughput",       0.04,   "$/GB",             "After free tier (10 GB/month)"),
    ("Pub/Sub — free tier",              10.0,    "GB/month",         "First 10 GB free per month"),
    ("GCS Standard Storage",             0.020,   "$/GB/month",       "asia-southeast1 region"),
    ("GCS Nearline Storage",             0.010,   "$/GB/month",       "For older snapshots (>30 days)"),
    ("GCS Write Ops (Class A)",          0.05,    "$/10,000 ops",     "PUT, POST to GCS"),
    ("GCS Read Ops (Class B)",           0.004,   "$/10,000 ops",     "GET from GCS"),
    ("GCS Egress (internet)",            0.08,    "$/GB",             "After 1 GB free/month"),
    ("Cloud Run — vCPU",                 0.000024,"$/vCPU-second",    "When processing requests"),
    ("Cloud Run — Memory",               0.0000025,"$/GB-second",     "When processing requests"),
    ("Cloud Run — Requests",             0.40,    "$/million req",    "After 2 million free/month"),
    ("GCE e2-micro (shared)",            5.48,    "$/month",          "0.25 vCPU, 1 GB RAM"),
    ("GCE e2-small (shared)",           10.96,    "$/month",          "0.5 vCPU, 2 GB RAM"),
    ("GCE e2-medium (shared)",          21.92,    "$/month",          "1 vCPU, 4 GB RAM"),
    ("GCE e2-standard-2",               48.92,    "$/month",          "2 vCPU, 8 GB RAM"),
]
for r, (label, val, unit, note) in enumerate(pricing, 25):
    style_label(ws.cell(r, 2), label)
    style_input(ws.cell(r, 3), val, USD if "$" in unit else NUM2)
    style_label(ws.cell(r, 4), unit)
    style_note(ws.cell(r, 5), note)

# Freeze panes
ws.freeze_panes = "B6"

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — USAGE & CAPACITY
# ══════════════════════════════════════════════════════════════════════════════
ws = sh_usage
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 40
ws.column_dimensions["C"].width = 20
ws.column_dimensions["D"].width = 15
ws.column_dimensions["E"].width = 35

ws.merge_cells("A1:E1")
style_header(ws["A1"], "Usage & Capacity Calculations", size=13)
ws.row_dimensions[1].height = 28

ws.merge_cells("A2:E2")
ws["A2"].value = "All values derived from '📥 Inputs' sheet"
ws["A2"].font = Font(italic=True, size=9, color="777777")
ws["A2"].alignment = align("center")

# Section headers
def section(ws, row, title, cols="A:E"):
    ws.merge_cells(f"A{row}:E{row}")
    style_section(ws[f"A{row}"], f"  {title}")
    ws.row_dimensions[row].height = 18

def row_calc(ws, row, label, formula, unit, note="", fmt=NUM):
    style_label(ws.cell(row, 2), label)
    style_formula(ws.cell(row, 3), formula, fmt)
    style_label(ws.cell(row, 4), unit)
    style_note(ws.cell(row, 5), note)

INP = "'📥 Inputs'"

section(ws, 4, "VOLUME — DETECTIONS & ALERTS")
calcs_vol = [
    ("Total Edge Devices",          f"={INP}!C6",                           "devices",      ""),
    ("Total Cameras",               f"={INP}!C8",                           "cameras",      "C6 × C7 from Inputs"),
    ("Detections per Camera/Hour",  f"={INP}!C9",                           "events/hr",    ""),
    ("Active Hours per Day",        f"={INP}!C10",                          "hrs/day",      ""),
    ("Alerts per Camera/Day",       f"=C9*C10",                             "alerts/day",   ""),
    ("Alerts per Edge/Day",         f"=C13*('{INP}!C7)",                    "alerts/day",   "cameras/edge × alerts/camera"),
    ("Total Alerts per Day",        f"=C6*C14",                             "alerts/day",   "all edges combined"),
    ("Total Alerts per Month",      f"=C15*{INP}!C19",                      "alerts/month", ""),
    ("Total Alerts per Year",       f"=C16*12",                             "alerts/year",  ""),
]

# Simpler direct formula approach
row = 5
labels_vol = [
    ("Total Edge Devices",       f"={INP}!C6",                                  "devices",     ""),
    ("Cameras per Edge",         f"={INP}!C7",                                  "cameras",     ""),
    ("Total Cameras",            f"=C5*C6",                                     "cameras",     ""),
    ("Detections / Camera / Hr", f"={INP}!C9",                                  "events/hr",   ""),
    ("Active Hours / Day",       f"={INP}!C10",                                 "hrs/day",     ""),
    ("Alerts / Camera / Day",    f"=C8*C9",                                     "alerts/day",  ""),
    ("Alerts / Edge / Day",      f"=C10*C6",                                    "alerts/day",  "sum per edge device"),
    ("Total Alerts / Day",       f"=C11*C5",                                    "alerts/day",  "all edges"),
    ("Total Alerts / Month",     f"=C12*{INP}!C19",                             "alerts/month",""),
    ("Total Alerts / Year",      f"=C13*12",                                    "alerts/year", ""),
]
for label, formula, unit, note in labels_vol:
    row_calc(ws, row, label, formula, unit, note)
    row += 1

section(ws, 16, "PUB/SUB THROUGHPUT")
row = 17
ps_rows = [
    ("Pub/Sub Msg Size",          f"={INP}!C21",                              "KB",          "from Inputs"),
    ("Messages / Month",          f"=C13",                                    "messages",    "1 msg per alert"),
    ("Data Volume / Month (KB)",  f"=C18*C17",                                "KB/month",    ""),
    ("Data Volume / Month (GB)",  f"=C19/1024/1024",                          "GB/month",    "", GB),
    ("Billable GB (after 10 free)",f"=MAX(0,C20-{INP}!C26)",                  "GB/month",    "first 10 GB free", GB),
]
for label, formula, unit, note, *rest in ps_rows:
    fmt = rest[0] if rest else NUM
    row_calc(ws, row, label, formula, unit, note, fmt)
    row += 1

section(ws, 23, "GCS STORAGE — SNAPSHOTS")
row = 24
gcs_rows = [
    ("Snapshot Size",             f"={INP}!C11",                             "KB",          "per alert"),
    ("Retention Period",          f"={INP}!C20",                             "days",        ""),
    ("Snapshots Stored",          f"=C13*(C25/{INP}!C19)",                   "snapshots",   "active retention window"),
    ("Storage Volume",            f"=C27*C24/1024/1024",                     "GB",          "total stored", GB),
    ("Write Ops / Month",         f"=C13",                                   "ops/month",   "1 write per alert"),
    ("Read Ops / Month",          f"=C13*0.3",                               "ops/month",   "est. 30% read rate for monitoring"),
]
for label, formula, unit, note, *rest in gcs_rows:
    fmt = rest[0] if rest else NUM
    row_calc(ws, row, label, formula, unit, note, fmt)
    row += 1

section(ws, 31, "SOCA-SERVICE — API CALL VOLUME")
row = 32
svc_rows = [
    ("Ingest Calls / Month",      f"=C13*{INP}!C16",                        "calls/month", "to soca-control /api/v1/ingest/"),
    ("Edge Poll Calls / Month",   f"=C5*(60/{INP}!C19)*1440*2",             "calls/month", "30s interval × 2 consumers"),
    ("Health Check Calls / Month",f"=C5*1440*{INP}!C19*2",                  "calls/month", "soca-control polls every 15s"),
    ("Total API Calls / Month",   f"=SUM(C32:C34)",                         "calls/month", ""),
]
for label, formula, unit, note in svc_rows:
    row_calc(ws, row, label, formula, unit, note)
    row += 1

ws.freeze_panes = "B5"

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — COST BREAKDOWN
# ══════════════════════════════════════════════════════════════════════════════
ws = sh_cost
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 38
ws.column_dimensions["C"].width = 20
ws.column_dimensions["D"].width = 20
ws.column_dimensions["E"].width = 20
ws.column_dimensions["F"].width = 30

ws.merge_cells("A1:F1")
style_header(ws["A1"], "Monthly Cost Breakdown", size=13)
ws.row_dimensions[1].height = 28

# Column headers
for col, hdr in enumerate(["Service / Component", "Monthly Volume", "Unit Price", "Monthly Cost", "Annual Cost", "Notes"], 2):
    c = ws.cell(3, col, hdr)
    c.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    c.fill = fill("37474F")
    c.alignment = align("center")
    c.border = border_all()
ws.row_dimensions[3].height = 18

USG = "'📊 Usage & Capacity'"
INP = "'📥 Inputs'"

def cost_row(ws, row, label, vol_formula, unit_formula, cost_formula, note=""):
    style_label(ws.cell(row, 2), label)
    style_formula(ws.cell(row, 3), vol_formula, NUM2)
    style_formula(ws.cell(row, 4), unit_formula, USD)
    style_formula(ws.cell(row, 5), cost_formula, USD)
    style_formula(ws.cell(row, 6), f"={chr(67+2)}{row}*12", USD)  # col E * 12
    style_note(ws.cell(row, 7) if False else ws.cell(row, 7), note)

row = 4
# PUB/SUB
ws.merge_cells(f"A{row}:F{row}")
style_section(ws[f"A{row}"], "  GOOGLE CLOUD PUB/SUB")
ws.row_dimensions[row].height = 18
row += 1

# PS row helper
def ps_row(r, label, vol_f, price_f, cost_f, note=""):
    style_label(ws.cell(r, 2), label)
    style_formula(ws.cell(r, 3), vol_f, NUM2)
    style_formula(ws.cell(r, 4), price_f, USD)
    style_formula(ws.cell(r, 5), cost_f, USD)
    style_formula(ws.cell(r, 6), f"=E{r}*12", USD)
    style_note(ws.cell(r, 7), note)

ps_row(row,   "Message Throughput (Billable GB)",
       f"={USG}!C21", f"={INP}!C25",
       f"={USG}!C21*{INP}!C25",
       "After 10 GB free tier deducted")
row += 1
ps_row(row,   "Pub/Sub — Free Tier Credit",
       "10", "—",
       f'=MIN(0, -{USG}!C20*{INP}!C25)',
       "First 10 GB/month free")
row += 1

ps_subtotal_row = row
ws.merge_cells(f"B{row}:D{row}")
style_total(ws.cell(row, 2), "Pub/Sub Subtotal")
style_total(ws.cell(row, 5), f"=E{row-2}+E{row-1}", USD)
style_total(ws.cell(row, 6), f"=F{row-2}+F{row-1}", USD)
row += 2

# GCS
ws.merge_cells(f"A{row}:F{row}")
style_section(ws[f"A{row}"], "  GOOGLE CLOUD STORAGE (GCS) — SNAPSHOTS")
ws.row_dimensions[row].height = 18
row += 1

gcs_start = row
ps_row(row, "Standard Storage",
       f"={USG}!C27", f"={INP}!C27",
       f"={USG}!C27*{INP}!C27",
       "Active retention window")
row += 1
ps_row(row, "Write Operations (Class A)",
       f"={USG}!C28/10000", f"={INP}!C29",
       f"=({USG}!C28/10000)*{INP}!C29",
       "Per 10,000 ops")
row += 1
ps_row(row, "Read Operations (Class B)",
       f"={USG}!C29/10000", f"={INP}!C30",
       f"=({USG}!C29/10000)*{INP}!C30",
       "Per 10,000 ops")
row += 1
ps_row(row, "Egress (est. 10% of snapshots)",
       f"={USG}!C27*0.1", f"={INP}!C31",
       f"={USG}!C27*0.1*{INP}!C31",
       "Estimated outbound traffic")
row += 1

gcs_end = row - 1
gcs_subtotal_row = row
ws.merge_cells(f"B{row}:D{row}")
style_total(ws.cell(row, 2), "GCS Subtotal")
style_total(ws.cell(row, 5), f"=SUM(E{gcs_start}:E{gcs_end})", USD)
style_total(ws.cell(row, 6), f"=SUM(F{gcs_start}:F{gcs_end})", USD)
row += 2

# INFRA
ws.merge_cells(f"A{row}:F{row}")
style_section(ws[f"A{row}"], "  CLOUD INFRASTRUCTURE — soca-control + soca-service")
ws.row_dimensions[row].height = 18
row += 1

infra_start = row
ps_row(row, "soca-control (VM / Cloud Run)",
       "1", f"='🖥 Infrastructure'!C8",
       f"='🖥 Infrastructure'!C8",
       "See Infrastructure sheet for sizing")
row += 1
ps_row(row, "soca-service (VM / Cloud Run)",
       "1", f"='🖥 Infrastructure'!C9",
       f"='🖥 Infrastructure'!C9",
       "See Infrastructure sheet for sizing")
row += 1
infra_end = row - 1

infra_subtotal_row = row
ws.merge_cells(f"B{row}:D{row}")
style_total(ws.cell(row, 2), "Infrastructure Subtotal")
style_total(ws.cell(row, 5), f"=SUM(E{infra_start}:E{infra_end})", USD)
style_total(ws.cell(row, 6), f"=SUM(F{infra_start}:F{infra_end})", USD)
row += 2

# GRAND TOTAL
ws.merge_cells(f"B{row}:D{row}")
c = ws.cell(row, 2)
c.value = "TOTAL MONTHLY CLOUD COST"
c.font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
c.fill = fill(C_HEADER_BG)
c.alignment = align("left")
c.border = border_all()

ws.cell(row, 5).value = f"=E{ps_subtotal_row}+E{gcs_subtotal_row}+E{infra_subtotal_row}"
ws.cell(row, 5).font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
ws.cell(row, 5).fill = fill(C_HEADER_BG)
ws.cell(row, 5).alignment = align("center")
ws.cell(row, 5).border = border_all()
ws.cell(row, 5).number_format = USD

ws.cell(row, 6).value = f"=F{ps_subtotal_row}+F{gcs_subtotal_row}+F{infra_subtotal_row}"
ws.cell(row, 6).font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
ws.cell(row, 6).fill = fill(C_HEADER_BG)
ws.cell(row, 6).alignment = align("center")
ws.cell(row, 6).border = border_all()
ws.cell(row, 6).number_format = USD
ws.row_dimensions[row].height = 24

ws.freeze_panes = "B4"

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — INFRASTRUCTURE SIZING
# ══════════════════════════════════════════════════════════════════════════════
ws = sh_infra
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 38
ws.column_dimensions["C"].width = 20
ws.column_dimensions["D"].width = 20
ws.column_dimensions["E"].width = 35

ws.merge_cells("A1:E1")
style_header(ws["A1"], "Infrastructure Sizing Guide", size=13)
ws.row_dimensions[1].height = 28

ws.merge_cells("A2:E2")
ws["A2"].value = "Select deployment type and instance. Costs flow to Cost Breakdown sheet."
ws["A2"].font = Font(italic=True, size=9, color="777777")
ws["A2"].alignment = align("center")

ws.merge_cells("A4:E4")
style_section(ws["A4"], "  CLOUD DEPLOYMENT SELECTION")
ws.row_dimensions[4].height = 18

for col, h in enumerate(["Component", "Selected Instance", "Monthly Cost", "Notes"], 2):
    c = ws.cell(5, col)
    c.value = h
    c.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    c.fill = fill("37474F")
    c.alignment = align("center")
    c.border = border_all()

infra_rows = [
    ("soca-control", "e2-small",  f"={INP}!C36", "Django + gunicorn, 2 workers"),
    ("soca-service", "e2-micro",  f"={INP}!C35", "FastAPI + uvicorn, lightweight"),
]
for r, (comp, inst, cost, note) in enumerate(infra_rows, 6):
    style_label(ws.cell(r, 2), comp)
    style_input(ws.cell(r, 3), inst)
    style_formula(ws.cell(r, 4), cost, USD)
    style_note(ws.cell(r, 5), note)

# Add validation dropdown for instance selection
dv = DataValidation(
    type="list",
    formula1='"e2-micro,e2-small,e2-medium,e2-standard-2,Cloud Run (auto)"',
    allow_blank=False,
    showDropDown=False,
)
ws.add_data_validation(dv)
dv.add("C6:C7")

ws.merge_cells("A9:E9")
style_section(ws["A9"], "  GCE VM SIZING REFERENCE")
ws.row_dimensions[9].height = 18

for col, h in enumerate(["Instance", "vCPU", "RAM (GB)", "Monthly Cost", "Recommended For"], 2):
    c = ws.cell(10, col)
    c.value = h
    c.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    c.fill = fill("37474F")
    c.alignment = align("center")
    c.border = border_all()

vm_specs = [
    ("e2-micro",       "0.25 (shared)", "1",  f"={INP}!C35", "soca-service only  (<50 edges)"),
    ("e2-small",       "0.5 (shared)",  "2",  f"={INP}!C36", "soca-control + service  (<100 edges)"),
    ("e2-medium",      "1 (shared)",    "4",  f"={INP}!C37", "Medium deployment  (100–500 edges)"),
    ("e2-standard-2",  "2",             "8",  f"={INP}!C38", "Large deployment  (500+ edges)"),
]
for r, (inst, cpu, ram, cost, rec) in enumerate(vm_specs, 11):
    style_label(ws.cell(r, 2), inst)
    ws.cell(r, 3).value = cpu; ws.cell(r, 3).alignment = align("center"); ws.cell(r, 3).border = border_all()
    ws.cell(r, 4).value = ram; ws.cell(r, 4).alignment = align("center"); ws.cell(r, 4).border = border_all()
    style_formula(ws.cell(r, 5), cost, USD)
    style_note(ws.cell(r, 6), rec)

# Edge sizing guide
ws.merge_cells("A16:E16")
style_section(ws["A16"], "  EDGE DEVICE SIZING (soca-dashboard + soca-engine)")
ws.row_dimensions[16].height = 18

ws.merge_cells("A17:E17")
ws["A17"].value = "Edge devices are on-premise — no cloud cost. Sizing depends on camera count and model type."
ws["A17"].font = Font(italic=True, size=9, color="555555")
ws["A17"].alignment = align("left", wrap=True)
ws.row_dimensions[17].height = 18

for col, h in enumerate(["Cameras", "Min RAM", "Min Storage", "GPU Required", "Notes"], 2):
    c = ws.cell(18, col)
    c.value = h
    c.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    c.fill = fill("37474F")
    c.alignment = align("center")
    c.border = border_all()

edge_specs = [
    ("1–4",    "4 GB",  "32 GB SSD",  "No",  "CPU inference, 1 job at a time"),
    ("4–8",    "8 GB",  "64 GB SSD",  "Optional", "CPU concurrent jobs or GPU inference"),
    ("8–16",   "16 GB", "128 GB SSD", "Recommended", "GPU required for real-time multi-stream"),
    ("16+",    "32 GB", "256 GB SSD", "Required", "Dedicated GPU server"),
]
for r, (cams, ram, stor, gpu, note) in enumerate(edge_specs, 19):
    for col, val in enumerate([cams, ram, stor, gpu], 2):
        ws.cell(r, col).value = val
        ws.cell(r, col).alignment = align("center")
        ws.cell(r, col).border = border_all()
        ws.cell(r, col).font = body_font()
    style_note(ws.cell(r, 6), note)

ws.freeze_panes = "B6"

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 5 — SUMMARY
# ══════════════════════════════════════════════════════════════════════════════
ws = sh_summary
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 35
ws.column_dimensions["C"].width = 22
ws.column_dimensions["D"].width = 22
ws.column_dimensions["E"].width = 22

ws.merge_cells("A1:E1")
style_header(ws["A1"], "SOCA Platform — Deployment Summary", size=14)
ws.row_dimensions[1].height = 32

ws.merge_cells("A2:E2")
ws["A2"].value = "Executive cost summary — all figures derived from Inputs sheet"
ws["A2"].font = Font(italic=True, size=10, color="555555")
ws["A2"].alignment = align("center")

INP = "'📥 Inputs'"
USG = "'📊 Usage & Capacity'"
CST = "'💰 Cost Breakdown'"

# ── Deployment config card ─────────────────────────────────────────────────────
ws.merge_cells("A4:E4")
style_section(ws["A4"], "  DEPLOYMENT CONFIGURATION")
ws.row_dimensions[4].height = 18

config_rows = [
    ("Edge Devices",     f"={INP}!C6",   "devices"),
    ("Total Cameras",    f"={INP}!C8",   "cameras"),
    ("Alerts / Month",   f"={USG}!C13",  "alerts"),
    ("Cloud Region",     f"={INP}!C15",  ""),
]
for r, (label, formula, unit) in enumerate(config_rows, 5):
    style_label(ws.cell(r, 2), label)
    style_formula(ws.cell(r, 3), formula, NUM)
    ws.cell(r, 4).value = unit
    ws.cell(r, 4).font = body_font()
    ws.cell(r, 4).border = border_all()
    ws.cell(r, 4).alignment = align("center")

# ── Cost summary ───────────────────────────────────────────────────────────────
ws.merge_cells("A10:E10")
style_section(ws["A10"], "  MONTHLY CLOUD COST SUMMARY")
ws.row_dimensions[10].height = 18

for col, h in enumerate(["Service", "Monthly Cost", "Annual Cost", "% of Total"], 2):
    c = ws.cell(11, col)
    c.value = h
    c.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    c.fill = fill("37474F")
    c.alignment = align("center")
    c.border = border_all()

# Cross-sheet references to cost breakdown subtotals
# We reference the subtotal rows from cost sheet (tracked as variables above)
cost_items = [
    ("Google Cloud Pub/Sub",       f"={CST}!E{ps_subtotal_row}",    f"={CST}!F{ps_subtotal_row}"),
    ("Google Cloud Storage (GCS)", f"={CST}!E{gcs_subtotal_row}",   f"={CST}!F{gcs_subtotal_row}"),
    ("Cloud Infrastructure",       f"={CST}!E{infra_subtotal_row}", f"={CST}!F{infra_subtotal_row}"),
]
summary_start = 12
gt = summary_start + len(cost_items)  # grand total row immediately after items (row 15)
for r, (label, monthly_f, annual_f) in enumerate(cost_items, summary_start):
    style_label(ws.cell(r, 2), label)
    style_formula(ws.cell(r, 3), monthly_f, USD)
    style_formula(ws.cell(r, 4), annual_f, USD)
    style_formula(ws.cell(r, 5), f"=C{r}/C{gt}", PCT)  # % of monthly total

# Grand total row
c = ws.cell(gt, 2)
c.value = "TOTAL MONTHLY COST"
c.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
c.fill = fill(C_HEADER_BG)
c.alignment = align("left")
c.border = border_all()

for col, formula, fmt in [
    (3, f"=SUM(C{summary_start}:C{gt-1})", USD),
    (4, f"=SUM(D{summary_start}:D{gt-1})", USD),
    (5, f"=100%", PCT),
]:
    c = ws.cell(gt, col)
    c.value = formula
    c.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    c.fill = fill(C_HEADER_BG)
    c.alignment = align("center")
    c.border = border_all()
    c.number_format = fmt
ws.row_dimensions[gt].height = 22

# ── Cost per camera ────────────────────────────────────────────────────────────
ws.merge_cells(f"A{gt+2}:E{gt+2}")
style_section(ws[f"A{gt+2}"], "  UNIT ECONOMICS")
ws.row_dimensions[gt+2].height = 18

unit_rows = [
    ("Cloud Cost per Camera / Month",   f"=C{gt}/{INP}!C8",          USD,                 "per camera",    "monthly cloud cost ÷ total cameras"),
    ("Cloud Cost per Edge / Month",     f"=C{gt}/{INP}!C6",           USD,                 "per edge",      "monthly cloud cost ÷ edge devices"),
    ("Cloud Cost per Alert",            f"=C{gt}/{USG}!C13",          '"$"#,##0.0000',     "per alert",     "monthly cloud cost ÷ total alerts/month"),
    ("GCS Storage per Camera (GB)",     f"={USG}!C27/{INP}!C8",       GB,                  "GB/camera",     "total GCS storage ÷ total cameras"),
    ("Pub/Sub Volume per Camera (MB)",  f"={USG}!C20/{INP}!C8*1024",  "#,##0.00 \"MB\"",   "MB/camera",     "monthly Pub/Sub volume ÷ total cameras"),
]
for r, (label, formula, fmt, unit, note) in enumerate(unit_rows, gt+3):
    style_label(ws.cell(r, 2), label)
    style_formula(ws.cell(r, 3), formula, fmt)
    c = ws.cell(r, 4)
    c.value = unit
    c.font = body_font()
    c.alignment = align("center")
    c.border = border_all()
    style_note(ws.cell(r, 5), note)

# ── Scaling tiers ──────────────────────────────────────────────────────────────
last = gt + 10
ws.merge_cells(f"A{last}:E{last}")
style_section(ws[f"A{last}"], "  SCALING TIERS REFERENCE")
ws.row_dimensions[last].height = 18

for col, h in enumerate(["Tier", "Edge Devices", "Total Cameras", "Est. Monthly Cost", "Recommended Infra"], 2):
    c = ws.cell(last+1, col)
    c.value = h
    c.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    c.fill = fill("37474F")
    c.alignment = align("center")
    c.border = border_all()

tiers = [
    ("Small",      "1–5",    "4–20",    "$5–15",    "e2-micro + e2-micro"),
    ("Medium",     "5–20",   "20–80",   "$15–50",   "e2-small + e2-micro"),
    ("Large",      "20–100", "80–400",  "$50–200",  "e2-medium + e2-small"),
    ("Enterprise", "100+",   "400+",    "$200+",    "e2-standard-2 + e2-small"),
]
for r, (tier, devs, cams, cost, infra) in enumerate(tiers, last+2):
    for col, val in enumerate([tier, devs, cams, cost, infra], 2):
        c = ws.cell(r, col)
        c.value = val
        c.font = body_font()
        c.alignment = align("center")
        c.border = border_all()
        if col == 5:
            c.fill = fill(C_GREEN_BG)
        elif col == 2:
            c.font = Font(name="Calibri", bold=True, size=10)

ws.freeze_panes = "B5"

# ══════════════════════════════════════════════════════════════════════════════
# SAVE
# ══════════════════════════════════════════════════════════════════════════════
out = "soca-capacity-cost-calculator.xlsx"
wb.save(out)
print(f"✓ Generated: {out}")
